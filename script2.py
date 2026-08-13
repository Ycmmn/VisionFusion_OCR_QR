# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
from pathlib import Path
import re
from datetime import datetime

# ============================================================================
# Section 1: Helper functions
# ============================================================================

def merge_cell_values(values):
    """Merge cell values, removing duplicates"""
    clean_values = [str(v).strip() for v in values if pd.notna(v) and str(v).strip() != '']
    if not clean_values:
        return ''
    seen = set()
    unique_values = []
    for val in clean_values:
        if val not in seen:
            seen.add(val)
            unique_values.append(val)
    if len(unique_values) == 1:
        return unique_values[0]
    return ' | '.join(unique_values)

def extract_base_name(col_name):
    """Extract the base name from a column (without the number)"""
    patterns = [
        r'^(.+?)\d+$',
        r'^(.+?)_\d+$',
        r'^(.+?)s\d+$',
    ]
    for pattern in patterns:
        match = re.match(pattern, col_name)
        if match:
            return match.group(1)
    return col_name

# ============================================================================
# Section 2: Cleanup functions
# ============================================================================

def clean_company_id(company_id_str):
    """Extract the main CompanyID from a merged string"""
    if pd.isna(company_id_str):
        return company_id_str
    pattern = r'COMP_UNKNOWN_[A-F0-9]+'
    match = re.search(pattern, str(company_id_str))
    if match:
        return match.group(0)
    parts = str(company_id_str).split('|')
    return parts[0].strip()

def extract_json_fields(text):
    """Extract name and position from JSON fields"""
    if pd.isna(text) or text == '':
        return None, None
    text = str(text)
    name_pattern = r"'name':\s*'([^']+)'"
    position_pattern = r"'position':\s*'([^']+)'"
    names = re.findall(name_pattern, text)
    positions = re.findall(position_pattern, text)
    name = ' | '.join(set(n for n in names if n and n != 'None')) if names else None
    position = ' | '.join(set(p for p in positions if p and p != 'None')) if positions else None
    return name, position

def remove_json_artifacts(text):
    """Remove JSON leftovers from text"""
    if pd.isna(text) or text == '':
        return text
    text = str(text)
    patterns = [
        r"\{'name':[^}]+\}",
        r"'name':\s*'[^']*'",
        r"'position':\s*'[^']*'",
        r"\|\s*'position':[^|]*",
        r"\s*\|\s*\|+\s*",
    ]
    for pattern in patterns:
        text = re.sub(pattern, '', text)
    text = text.strip()
    text = re.sub(r'^\|+\s*|\s*\|+$', '', text)
    text = re.sub(r'\s*\|\s*\|\s*', ' | ', text)
    return text.strip() if text.strip() else None

# ============================================================================
# Section 3: Merging numbered columns
# ============================================================================

def find_numbered_groups(columns):
    """Find groups of numbered columns"""
    from collections import defaultdict
    groups = defaultdict(list)
    for col in columns:
        base_name = extract_base_name(col.lower())
        groups[base_name].append(col)
    result = {k: v for k, v in groups.items() if len(v) > 1}
    return result

def merge_numbered_columns(df, verbose=True):
    """Merge numbered columns"""
    if verbose:
        print("در حال شناسایی ستون‌های شماره‌دار...")
    groups = find_numbered_groups(df.columns.tolist())
    if not groups:
        if verbose:
            print("گروه شماره‌داری پیدا نشد")
        return df, 0
    if verbose:
        print(f"{len(groups)} گروه شماره‌دار پیدا شد")
    merged_count = 0
    for base_name, cols in groups.items():
        cols_sorted = sorted(cols, key=lambda x: (len(x), x))
        main_col = cols_sorted[0]
        other_cols = cols_sorted[1:]
        for idx in df.index:
            values = []
            for col in cols_sorted:
                val = df.at[idx, col]
                if pd.notna(val) and str(val).strip() != '':
                    values.append(str(val).strip())
            unique_values = list(dict.fromkeys(values))
            if unique_values:
                if len(unique_values) == 1:
                    df.at[idx, main_col] = unique_values[0]
                else:
                    df.at[idx, main_col] = ' | '.join(unique_values)
        df.drop(columns=other_cols, inplace=True)
        merged_count += len(other_cols)
    if verbose:
        print(f"{merged_count} ستون شماره‌دار حذف شد")
    return df, merged_count

# ============================================================================
# Section 4: Merging duplicate columns
# ============================================================================

def merge_duplicate_columns(df, verbose=True):
    """Merge duplicate columns (case-insensitive)"""
    if verbose:
        print("در حال بررسی ستون‌های تکراری...")
    column_map = {}
    for col in df.columns:
        col_lower = col.lower()
        if col_lower not in column_map:
            column_map[col_lower] = []
        column_map[col_lower].append(col)
    duplicated_groups = {k: v for k, v in column_map.items() if len(v) > 1}
    if not duplicated_groups:
        if verbose:
            print("ستون تکراری پیدا نشد")
        return df, 0
    if verbose:
        print(f"{len(duplicated_groups)} گروه ستون تکراری پیدا شد")
    merged_count = 0
    new_df = pd.DataFrame()
    processed_cols = set()
    for col in df.columns:
        if col in processed_cols:
            continue
        col_lower = col.lower()
        if col_lower not in duplicated_groups:
            new_df[col] = df[col]
            processed_cols.add(col)
        else:
            all_same_cols = column_map[col_lower]
            if all_same_cols[0] in processed_cols:
                continue
            merged_col = []
            for idx in df.index:
                values = [df[c].iloc[idx] for c in all_same_cols]
                merged_col.append(merge_cell_values(values))
            final_name = all_same_cols[0]
            new_df[final_name] = merged_col
            for c in all_same_cols:
                processed_cols.add(c)
            merged_count += (len(all_same_cols) - 1)
    if verbose:
        print(f"{merged_count} ستون تکراری حذف شد")
    return new_df, merged_count

# ============================================================================
# Section 5: Merging bilingual columns (EN | FA into one column)
# ============================================================================

def find_bilingual_pairs(columns):
    """Find English/Persian column pairs"""
    pairs = []
    processed = set()

    # explicit pairs: column + same column + FA
    fa_suffixes = ['FA', '_translated']

    for col in columns:
        if col in processed:
            continue
        for fa_suffix in fa_suffixes:
            fa_col = col + fa_suffix
            if fa_col in columns and fa_col not in processed:
                pairs.append((col, fa_col))
                processed.add(col)
                processed.add(fa_col)
                break

        # columns ending with EN
        if col not in processed and col.endswith('EN'):
            base = col[:-2]
            fa_col = base + 'FA'
            if fa_col in columns and fa_col not in processed:
                pairs.append((col, fa_col))
                processed.add(col)
                processed.add(fa_col)

    return pairs

def merge_bilingual_columns(df, verbose=True):
    """
    Merge English/Persian columns:
    EN column and FA column -> one column with value 'English value | Persian value'
    The FA column is removed and the EN column is kept
    """
    if verbose:
        print("در حال شناسایی ستون‌های دوزبانه...")

    pairs = find_bilingual_pairs(df.columns.tolist())

    if not pairs:
        if verbose:
            print("جفت دوزبانه‌ای پیدا نشد")
        return df, 0

    if verbose:
        print(f"{len(pairs)} جفت دوزبانه پیدا شد:")
        for en, fa in pairs:
            print(f"   {en} + {fa} → {en}")

    merged_count = 0

    for en_col, fa_col in pairs:
        if en_col not in df.columns or fa_col not in df.columns:
            continue

        for idx in df.index:
            en_val = df.at[idx, en_col]
            fa_val = df.at[idx, fa_col]

            values = []
            if pd.notna(en_val) and str(en_val).strip() not in ('', 'nan', 'None'):
                values.append(str(en_val).strip())
            if pd.notna(fa_val) and str(fa_val).strip() not in ('', 'nan', 'None'):
                fa_str = str(fa_val).strip()
                # avoid duplication (if the Persian value is the same as the English one)
                if fa_str not in values:
                    values.append(fa_str)

            if values:
                df.at[idx, en_col] = ' | '.join(values)
            else:
                df.at[idx, en_col] = None

        df.drop(columns=[fa_col], inplace=True)
        merged_count += 1

    if verbose:
        print(f"{merged_count} جفت دوزبانه ادغام شد")

    return df, merged_count

# ============================================================================
# Section 6: Merging specific columns
# ============================================================================

def merge_specific_columns(df, verbose=True):
    """Merge specific columns that need to be combined"""
    if verbose:
        print("در حال ادغام ستون‌های خاص...")
    merged_count = 0
    merge_groups = [
        ('fax', ['faxes', 'Fax']),
        ('phones', ['phone1', 'Phone1', 'Phone', 'phone']),
        ('urls', ['url', 'Website', 'website', 'URL']),
        ('company_names', ['CompanyNameFA_translated', 'CompanyNameEN']),
        ('emails', ['Email', 'OtherEmails']),
    ]
    for main_col_pattern, other_patterns in merge_groups:
        main_col = None
        for col in df.columns:
            if col.lower() == main_col_pattern.lower():
                main_col = col
                break
        other_cols = []
        for pattern in other_patterns:
            for col in df.columns:
                if col.lower() == pattern.lower() and col != main_col:
                    other_cols.append(col)
                    break
        if main_col and other_cols:
            if verbose:
                print(f"ادغام '{main_col}' با {other_cols}")
            for idx in df.index:
                values = []
                main_val = df.at[idx, main_col]
                if pd.notna(main_val) and str(main_val).strip() != '':
                    values.append(str(main_val).strip())
                for col in other_cols:
                    val = df.at[idx, col]
                    if pd.notna(val) and str(val).strip() != '':
                        values.append(str(val).strip())
                unique_values = list(dict.fromkeys(values))
                if unique_values:
                    df.at[idx, main_col] = ' | '.join(unique_values) if len(unique_values) > 1 else unique_values[0]
            df.drop(columns=other_cols, inplace=True)
            merged_count += len(other_cols)
        elif not main_col and other_cols:
            main_col = other_cols[0]
            remaining_cols = other_cols[1:]
            if remaining_cols:
                for idx in df.index:
                    values = []
                    for col in other_cols:
                        val = df.at[idx, col]
                        if pd.notna(val) and str(val).strip() != '':
                            values.append(str(val).strip())
                    unique_values = list(dict.fromkeys(values))
                    if unique_values:
                        df.at[idx, main_col] = ' | '.join(unique_values) if len(unique_values) > 1 else unique_values[0]
                df.drop(columns=remaining_cols, inplace=True)
                merged_count += len(remaining_cols)
    if verbose:
        print(f"جمع {merged_count} ستون خاص ادغام شد" if merged_count > 0 else "ستون خاصی برای ادغام پیدا نشد")
    return df, merged_count

# ============================================================================
# Merging rows
# ============================================================================

def merge_rows_by_company_id(df, company_id_col=None, verbose=True):
    """Merge rows with the same company_id"""
    if company_id_col is None:
        for col in df.columns:
            if 'company' in col.lower() and 'id' in col.lower():
                company_id_col = col
                break
    if company_id_col is None:
        company_id_col = df.columns[0]
        if verbose:
            print(f"از ستون اول استفاده می‌شود: {company_id_col}")
    else:
        if verbose:
            print(f"ستون company_id: {company_id_col}")
    grouped = df.groupby(company_id_col, dropna=False)
    if verbose:
        print(f"در حال ادغام {len(grouped)} گروه...")
    merged_rows = []
    for company_id, group in grouped:
        if len(group) == 1:
            merged_rows.append(group.iloc[0].to_dict())
        else:
            merged_row = {}
            for col in df.columns:
                if col == company_id_col:
                    merged_row[col] = company_id
                else:
                    values = group[col].tolist()
                    merged_row[col] = merge_cell_values(values)
            merged_rows.append(merged_row)
    return pd.DataFrame(merged_rows)

# ============================================================================
# Section 7: Final cleanup
# ============================================================================

def clean_data(df, verbose=True):
    """Final cleanup of the data"""
    if verbose:
        print("تمیز کردن CompanyID...")
    if 'CompanyID' in df.columns:
        df['CompanyID'] = df['CompanyID'].apply(clean_company_id)
        if verbose:
            print("CompanyID تمیز شد")
    if verbose:
        print("پاکسازی عمومی...")
    skip_columns = ['Name', 'Position', 'ContactName', 'name', 'position']
    cleaned_count = 0
    for col in df.columns:
        if col in skip_columns or col == 'CompanyID':
            continue
        for idx in df.index:
            val = df.at[idx, col]
            if pd.notna(val) and isinstance(val, str):
                if ('{' in val or "'name':" in val or "'position':" in val) and col not in ['file_name']:
                    cleaned = remove_json_artifacts(val)
                    if cleaned != val and cleaned:
                        df.at[idx, col] = cleaned
                        cleaned_count += 1
    if verbose and cleaned_count > 0:
        print(f"{cleaned_count} سلول تمیز شد")
    return df

# ============================================================================
# Section 8: Standardizing URL and Phone
# ============================================================================

def standardize_url(url):
    if pd.isna(url) or str(url).strip() == '':
        return None
    url = str(url).strip().replace(' ', '')
    if not url.startswith('http://') and not url.startswith('https://'):
        url = 'https://' + url
    if url.startswith('http://'):
        url = url.replace('http://', 'https://', 1)
    if url.startswith('https://') and not url.startswith('https://www.'):
        domain_part = url.replace('https://', '')
        if domain_part.count('.') <= 1:
            url = 'https://www.' + domain_part
    return url

def find_duplicate_urls(urls_list):
    from urllib.parse import urlparse
    url_map = {}
    for url in urls_list:
        if not url:
            continue
        try:
            parsed = urlparse(url)
            domain = parsed.netloc.lower().replace('www.', '')
            if domain in url_map:
                if url.startswith('https://www.') and not url_map[domain].startswith('https://www.'):
                    url_map[domain] = url
            else:
                url_map[domain] = url
        except:
            url_map[url] = url
    return list(url_map.values())

def remove_duplicates_from_cell(value, is_url=False):
    if pd.isna(value) or str(value).strip() == '':
        return None
    items = [item.strip() for item in str(value).split('|')]
    if is_url:
        standardized = [standardize_url(i) for i in items if standardize_url(i)]
        unique_urls = find_duplicate_urls(standardized)
        return ' | '.join(unique_urls) if unique_urls else None
    else:
        normalized = []
        seen = set()
        for item in items:
            item_compare = item.replace(' ', '').replace('-', '').replace('(', '').replace(')', '')
            if item_compare and item_compare not in seen:
                seen.add(item_compare)
                normalized.append(item)
        return ' | '.join(normalized) if normalized else None

def clean_urls_and_phones(df, verbose=True):
    if verbose:
        print("تمیز کردن URLs و تلفن‌ها...")
    cleaned_count = 0
    if 'urls' in df.columns:
        for idx in df.index:
            val = df.at[idx, 'urls']
            if pd.notna(val) and str(val).strip() != '':
                cleaned = remove_duplicates_from_cell(val, is_url=True)
                if cleaned != val:
                    df.at[idx, 'urls'] = cleaned
                    cleaned_count += 1
    phone_count = 0
    if 'phones' in df.columns:
        for idx in df.index:
            val = df.at[idx, 'phones']
            if pd.notna(val) and str(val).strip() != '':
                cleaned = remove_duplicates_from_cell(val, is_url=False)
                if cleaned and cleaned != val:
                    df.at[idx, 'phones'] = cleaned
                    phone_count += 1
    if verbose:
        total = cleaned_count + phone_count
        print(f"جمع {total} سلول تمیز شد" if total > 0 else "تکراری پیدا نشد")
    return df

# ============================================================================
# Main function
# ============================================================================

def process_company_data(
    input_file,
    output_file=None,
    keep_empty_columns=True,
    company_id_col=None,
    verbose=True
):
    """Full processing of company data"""
    if verbose:
        print("="*60)
        print("شروع پردازش کامل داده‌های شرکت‌ها")
        print("="*60)
        print(f"در حال خواندن: {input_file}")

    if str(input_file).endswith('.csv'):
        df = pd.read_csv(input_file, encoding='utf-8-sig')
    else:
        df = pd.read_excel(input_file)

    initial_rows = len(df)
    initial_cols = len(df.columns)

    if verbose:
        print(f"{initial_rows} سطر، {initial_cols} ستون")

    if verbose: print("\n" + "="*60 + "\nمرحله 1: ادغام ستون‌های شماره‌دار\n" + "="*60)
    df, numbered_merged = merge_numbered_columns(df, verbose=verbose)

    if verbose: print("\n" + "="*60 + "\nمرحله 2: ادغام ستون‌های تکراری\n" + "="*60)
    df, duplicate_merged = merge_duplicate_columns(df, verbose=verbose)

    if verbose: print("\n" + "="*60 + "\nمرحله 3: ادغام ستون‌های دوزبانه (EN | FA)\n" + "="*60)
    df, bilingual_merged = merge_bilingual_columns(df, verbose=verbose)

    if verbose: print("\n" + "="*60 + "\nمرحله 3.5: ادغام ستون‌های خاص\n" + "="*60)
    df, specific_merged = merge_specific_columns(df, verbose=verbose)

    if verbose: print("\n" + "="*60 + "\nمرحله 4: ادغام سطرها\n" + "="*60)
    df = merge_rows_by_company_id(df, company_id_col, verbose=verbose)
    rows_after_merge = len(df)

    if verbose: print("\n" + "="*60 + "\nمرحله 5: تمیزکاری نهایی\n" + "="*60)
    df = clean_data(df, verbose=verbose)

    if verbose: print("\n" + "="*60 + "\nمرحله 5.5: استاندارد‌سازی URLs و Phones\n" + "="*60)
    df = clean_urls_and_phones(df, verbose=verbose)

    if verbose: print("\n" + "="*60 + "\nمرحله 6: مدیریت ستون‌های خالی\n" + "="*60)
    empty_cols_removed = 0
    if not keep_empty_columns:
        empty_before = len(df.columns)
        df = df.dropna(axis=1, how='all')
        empty_cols_removed = empty_before - len(df.columns)
        if verbose and empty_cols_removed > 0:
            print(f"{empty_cols_removed} ستون خالی حذف شد")
    else:
        if verbose:
            print("تمام ستون‌ها حفظ شدند")

    if output_file is None:
        input_path = Path(input_file)
        output_file = input_path.parent / f"{input_path.stem}_processed{input_path.suffix}"
        counter = 1
        while output_file.exists():
            try:
                with open(output_file, 'a'):
                    pass
                break
            except PermissionError:
                output_file = input_path.parent / f"{input_path.stem}_processed_{counter}{input_path.suffix}"
                counter += 1

    if verbose:
        print(f"در حال ذخیره: {output_file}")

    try:
        if str(output_file).endswith('.csv'):
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
        else:
            df.to_excel(output_file, index=False, engine='openpyxl')
        if verbose:
            print("فایل ذخیره شد")
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        input_path = Path(input_file)
        output_file = input_path.parent / f"{input_path.stem}_processed_{timestamp}{input_path.suffix}"
        if verbose:
            print(f"فایل باز است، ذخیره با نام: {output_file}")
        if str(output_file).endswith('.csv'):
            df.to_csv(output_file, index=False, encoding='utf-8-sig')
        else:
            df.to_excel(output_file, index=False, engine='openpyxl')

    if verbose:
        print("\n" + "="*60)
        print("خلاصه نهایی")
        print("="*60)
        print(f"سطرهای ورودی: {initial_rows}")
        print(f"سطرهای خروجی: {rows_after_merge}")
        print(f"ستون‌های ورودی: {initial_cols}")
        print(f"ستون‌های خروجی: {len(df.columns)}")
        total_merged = numbered_merged + duplicate_merged + bilingual_merged + specific_merged
        if total_merged > 0:
            print(f"\nستون‌های ادغام شده:")
            if numbered_merged > 0: print(f"   - شماره‌دار: {numbered_merged}")
            if duplicate_merged > 0: print(f"   - تکراری: {duplicate_merged}")
            if bilingual_merged > 0: print(f"   - دوزبانه (EN|FA): {bilingual_merged}")
            if specific_merged > 0: print(f"   - خاص: {specific_merged}")
            print(f"   - جمع: {total_merged}")
        print("="*60)
        print(f"فایل: {output_file}")
        print("="*60)

    return df

# ============================================================================
# Run
# ============================================================================

if __name__ == "__main__":
    input_file = "Exhibition_QC_Data - Sheet1.csv"
    try:
        df_result = process_company_data(
            input_file=input_file,
            keep_empty_columns=True,
            verbose=True
        )
        print("\nپردازش با موفقیت کامل شد!")
    except FileNotFoundError:
        print(f"خطا: فایل '{input_file}' پیدا نشد!")
    except Exception as e:
        print(f"خطا: {str(e)}")
        import traceback
        traceback.print_exc()

# ============================================================================
# Wrapper for use in the Pipeline
# ============================================================================



def highlight_duplicate_companies(output_path):
    """
    Color rows that share the same CompanyName
    Each group gets a different color
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    
    COLORS = [
        "FFE699",  # yellow
        "BDD7EE",  # blue
        "C6EFCE",  # green
        "FCE4D6",  # orange
        "E2EFDA",  # light green
        "D9D2E9",  # purple
        "FCE4D6",  # pink
        "DDEBF7",  # light blue
    ]
    
    wb = load_workbook(output_path)
    ws = wb.active
    
    # find the CompanyName column
    header_row = [cell.value for cell in ws[1]]
    
    company_col_idx = None
    for i, header in enumerate(header_row):
        if header and 'companyname' in str(header).lower().replace('_', '').replace(' ', ''):
            company_col_idx = i + 1  # openpyxl starts from 1
            break
    
    if company_col_idx is None:
        print("    CompanyName column not found for highlighting")
        wb.close()
        return
    
    print(f"    Found CompanyName at column {company_col_idx}")
    
    # collect the CompanyName value of each row
    company_rows = {}  # CompanyName -> [row numbers]
    
    for row in ws.iter_rows(min_row=2):
        cell = row[company_col_idx - 1]
        company_name = str(cell.value).strip() if cell.value else ""
        
        if not company_name or company_name in ['', 'None', 'nan']:
            continue
        
        # normalize for comparison
        normalized = company_name.lower().strip()
        
        if normalized not in company_rows:
            company_rows[normalized] = []
        company_rows[normalized].append(cell.row)
    
    # only companies that have more than one row
    duplicate_companies = {k: v for k, v in company_rows.items() if len(v) > 1}
    
    print(f"    Found {len(duplicate_companies)} duplicate companies")
    
    # coloring
    color_idx = 0
    colored_count = 0
    
    for company_name, rows in duplicate_companies.items():
        color_hex = COLORS[color_idx % len(COLORS)]
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        
        for row_num in rows:
            for cell in ws[row_num]:
                cell.fill = fill
            colored_count += 1
        
        color_idx += 1
    
    wb.save(output_path)
    wb.close()
    
    print(f"   Colored {colored_count} rows across {len(duplicate_companies)} duplicate companies")






def script2_process_file(input_path, output_path):
    df = process_company_data(
        input_file=input_path,
        output_file=output_path,
        keep_empty_columns=True,
        company_id_col=None,
        verbose=False
    )
    
    # ==========  Highlight duplicate companies ==========
    try:
        highlight_duplicate_companies(output_path)
    except Exception as e:
        print(f"    Highlighting failed: {e}")
    
    return df